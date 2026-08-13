from flask import Flask, request, jsonify, make_response
from flask_cors import CORS
from openpyxl import load_workbook
import os, json, datetime, difflib, threading, time, gc
try:
    import psycopg2
    import psycopg2.extras
    HAS_PSYCOPG2 = True
except ImportError:
    HAS_PSYCOPG2 = False
import pytz

app = Flask(__name__)
CORS(app)

DATA_FILE          = "mis_data.json"
AUDIT_FILE         = "audit.json"
UNIQUE_FILE        = "unique_set.json"
HANDOFF_MOBILE_FILE= "handoff_mobile_set.json"
BLENDED_CID_FILE   = "blended_cid_set.json"
MANUAL_CID_FILE    = "manual_cid_set.json"
IST = pytz.timezone("Asia/Kolkata")

# ── PostgreSQL KV Store ───────────────────────────────────────────────────────
_db_conn = None

def get_db():
    global _db_conn
    if not HAS_PSYCOPG2:
        return None
    db_url  = os.environ.get("DATABASE_URL")
    db_host = os.environ.get("DB_HOST")
    try:
        if _db_conn is None or _db_conn.closed:
            if db_url:
                _db_conn = psycopg2.connect(db_url, connect_timeout=10)
            elif db_host:
                _db_conn = psycopg2.connect(
                    host=db_host,
                    database=os.environ.get("DB_NAME", "postgres"),
                    user=os.environ.get("DB_USER", "postgres"),
                    password=os.environ.get("DB_PASSWORD", ""),
                    port=int(os.environ.get("DB_PORT", 5432)),
                    sslmode="require",
                    connect_timeout=10,
                )
            else:
                return None
            _db_conn.autocommit = True
            with _db_conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS carmaone_kv (
                        key TEXT PRIMARY KEY,
                        value TEXT NOT NULL,
                        updated_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
        return _db_conn
    except Exception as e:
        print(f"DB connection error: {e}")
        return None

def db_get(key, default=None):
    global _db_conn
    for attempt in range(2):
        conn = get_db()
        if conn is None:
            break
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT value FROM carmaone_kv WHERE key = %s", (key,))
                row = cur.fetchone()
                return json.loads(row[0]) if row else default
        except Exception as e:
            print(f"db_get error {key} attempt {attempt+1}: {e}")
            _db_conn = None
    data = load_json(key + ".json")
    return data if data is not None else default

def db_set(key, value):
    global _db_conn
    for attempt in range(2):
        conn = get_db()
        if conn is None:
            break
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO carmaone_kv (key, value, updated_at)
                    VALUES (%s, %s, NOW())
                    ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
                """, (key, json.dumps(value, ensure_ascii=False)))
            return
        except Exception as e:
            print(f"db_set error {key} attempt {attempt+1}: {e}")
            _db_conn = None
    save_json(key + ".json", value)

def db_get_set(key):
    data = db_get(key)
    return set(data) if isinstance(data, list) else set()

def db_set_set(key, s):
    db_set(key, list(s))

def db_get_monthly_sets(prefix):
    conn = get_db()
    result = {}
    if conn:
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT key, value FROM carmaone_kv WHERE key LIKE %s",
                            (prefix + ":%",))
                for row in cur.fetchall():
                    ym = row[0].split(":", 1)[1]
                    result[ym] = set(json.loads(row[1]))
        except Exception as e:
            print(f"db_get_monthly_sets error: {e}")
    return result

def db_set_monthly_sets(prefix, monthly_sets):
    for ym, s in monthly_sets.items():
        db_set_set(f"{prefix}:{ym}", s)

def _supabase_keepalive():
    while True:
        time.sleep(20 * 60)
        try:
            conn = get_db()
            if conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT 1")
                print("Supabase keepalive OK")
        except Exception as e:
            print(f"Supabase keepalive error: {e}")
            global _db_conn
            _db_conn = None

threading.Thread(target=_supabase_keepalive, daemon=True).start()

PINS = {"1212": "Tarun", "7890": "Dipanshu"}

KNOWN_STATUSES = [
    "Disbursed", "Rejected", "Sanctioned", "Case Dropped",
    "Login Pending", "Login", "Hold", "Credit Review"
]
NOT_CONNECTED_DISPOS = {"NO ANSWER", "NUMBER BUSY", "SWITCHED OFF", "DUPLICATE LEAD", "BLANK"}
VALID_CALLMODES = {
    "predictive", "predictive-blended", "manual",
    "callback", "redial", "outbound", "progressive"
}
VALID_HANDOFF_DISPOS = {"INTERESTED", "CALL BACK"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def now_ist():
    return datetime.datetime.now(IST).strftime("%d %b %Y, %I:%M %p IST")

def load_json(path):
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return None

def save_json(path, data):
    with open(path, "w") as f:
        json.dump(data, f, ensure_ascii=False)

def load_unique_set():
    return db_get_set("unique_set")

def save_unique_set(s):
    db_set_set("unique_set", s)

def load_set(key):
    return db_get_set(key)

def save_set(key, s):
    db_set_set(key, s)

def normalize_mobile(raw):
    s = str(raw).strip().replace(' ', '').replace('-', '')
    if s.startswith('+91'):
        s = s[3:]
    elif s.startswith('91') and len(s) == 12:
        s = s[2:]
    return s

def excel_to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        try:
            d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(val))
            return d.strftime("%Y-%m-%d")
        except Exception:
            return None
    if isinstance(val, str):
        s = val.strip()
        if ' ' in s and ':' in s:
            s = s.split(' ')[0]
        for fmt in [
            "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
            "%d %b %Y", "%d %B %Y",
            "%d-%b-%Y", "%d-%B-%Y",
            "%m/%d/%Y", "%Y/%m/%d",
            "%d/%m/%y", "%d-%m-%y",
        ]:
            try:
                return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                continue
    return None

def date_label(iso):
    if not iso:
        return ""
    try:
        return datetime.datetime.strptime(iso, "%Y-%m-%d").strftime("%-d %b %Y")
    except Exception:
        return iso

def normalize_status(raw):
    if not raw:
        return ("Unknown", False)
    s = str(raw).strip()
    for k in KNOWN_STATUSES:
        if k.lower() == s.lower():
            return (k, False)
    matches = difflib.get_close_matches(s, KNOWN_STATUSES, n=1, cutoff=0.6)
    if matches:
        return (matches[0], True)
    # Never pass an unrecognised value through: it would render as a phantom
    # status card on the dashboard. validate_pipeline rejects these upfront —
    # this is the backstop.
    return ("Unknown", False)

def sf(v, d=0.0):
    try:
        return float(v) if v is not None else d
    except Exception:
        return d

def si(v, d=0):
    try:
        return int(float(v)) if v is not None else d
    except Exception:
        return d

def ss(v):
    return str(v).strip() if v is not None else ""

def normalize_agent(name):
    if not name:
        return ""
    parts = str(name).strip().split()
    deduped = [parts[0]] if parts else []
    for p in parts[1:]:
        if p.lower() != deduped[-1].lower():
            deduped.append(p)
    return " ".join(deduped).title()


# ── Sheet reader ──────────────────────────────────────────────────────────────

def _norm(n):
    return n.lower().replace(" ","").replace("_","").replace("-","")

def find_sheet_name(wb, candidates):
    for c in candidates:
        if c in wb.sheetnames:
            return c
    for c in candidates:
        for s in wb.sheetnames:
            if s.lower() == c.lower():
                return s
    for c in candidates:
        nc = _norm(c)
        for s in wb.sheetnames:
            if _norm(s) == nc:
                return s
    for c in candidates:
        for s in wb.sheetnames:
            if _norm(c) in _norm(s) or _norm(s) in _norm(c):
                return s
    return None

def read_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header, header_idx = None, 0
    for i, r in enumerate(rows):
        if any(c is not None for c in r):
            header = [ss(c) for c in r]
            header_idx = i
            break
    if not header:
        return []
    result = []
    for r in rows[header_idx + 1:]:
        if all(c is None for c in r):
            continue
        row = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        result.append(row)
    return result

def stream_sheet(wb, sheet_name, cols=None):
    """Generator: yield row dicts one at a time — never materialises full list."""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    header = None
    col_idx = None
    for r in ws.iter_rows(values_only=True):
        if header is None:
            if any(c is not None for c in r):
                header = [ss(c) for c in r]
                col_idx = {h: i for i, h in enumerate(header)
                           if not cols or h in cols}
            continue
        if all(c is None for c in r):
            continue
        yield {h: r[i] for h, i in col_idx.items() if i < len(r)}

def normalize_jarvis_rows(rows):
    """Convert Jarvis_Dump rows to Flexy-compatible format."""
    normalized = []
    for r in rows:
        direction = ss(r.get("Direction", "")).lower()
        if direction == "inbound":
            continue
        agent = normalize_agent(ss(r.get("Agent Name", "")))
        if not agent:
            continue
        cid    = ss(str(r.get("Number", "")))
        status = ss(r.get("Status", "")).lower()
        dispo  = ss(r.get("AI_Tag", "")) or "Other"
        dur_s  = sf(r.get("Duration (s)", 0))
        date   = r.get("Started At")
        dialed = "connected" if status == "answered" else "not_connected"
        # Only map productive dispositions for answered calls
        if status == "answered":
            if dispo.lower() == "interested":
                primary_dispo = "Interested"
            elif dispo.lower() in ("callback", "call back"):
                primary_dispo = "Call Back"
            else:
                primary_dispo = "Not Interested"
        else:
            primary_dispo = "Not Answered"
        normalized.append({
            "customer_cid":  normalize_mobile(cid),
            "user":          agent,
            "callmode":      "manual",
            "init_time":     date,
            "talk_time":     dur_s / 86400.0,
            "primary_dispo": primary_dispo,
            "dialed_status": dialed,
            "_source":       "jarvis",
        })
    return normalized


# ── Validators ────────────────────────────────────────────────────────────────

def _e(sheet, row, msg):
    return {"sheet": sheet, "row": str(row), "msg": msg}

def _w(sheet, row, msg):
    return {"sheet": sheet, "row": str(row), "msg": msg}

def validate_ai_dump(rows):
    errors, warnings = [], []
    required = ["#", "Call Date", "Batch", "L1 Disposition", "Call Duration"]
    sheet = "AI Dump"
    if not rows:
        return [_e(sheet, "-", "Sheet is empty or not found")], []
    for col in required:
        if col not in rows[0]:
            errors.append(_e(sheet, "Header", f"Missing required column: '{col}'"))
    if errors:
        return errors, warnings
    for i, r in enumerate(rows, 2):
        rid  = ss(r.get("#", "")) or str(i)
        date = excel_to_date(r.get("Call Date"))
        dur  = sf(r.get("Call Duration", 0))
        bat  = ss(r.get("Batch", ""))
        if not bat:
            errors.append(_e(sheet, rid, "Batch is empty"))
        if date is None:
            errors.append(_e(sheet, rid, "Invalid or missing Call Date"))
        if dur < 0:
            errors.append(_e(sheet, rid, f"Call Duration is negative ({dur})"))
    return errors, warnings

def validate_handoff(rows):
    errors, warnings = [], []
    required = ["#", "Call Date", "Batch", "L1 Disposition"]
    sheet = "Handoff Leads"
    if not rows:
        return [_e(sheet, "-", "Sheet is empty or not found")], []
    for col in required:
        if col not in rows[0]:
            errors.append(_e(sheet, "Header", f"Missing required column: '{col}'"))
    if errors:
        return errors, warnings
    for i, r in enumerate(rows, 2):
        rid   = ss(r.get("#", "")) or str(i)
        date  = excel_to_date(r.get("Call Date"))
        dispo = ss(r.get("L1 Disposition", "")).upper()
        if date is None:
            errors.append(_e(sheet, rid, "Invalid or missing Call Date"))
        if dispo not in VALID_HANDOFF_DISPOS:
            errors.append(_e(sheet, rid,
                f"Invalid disposition '{dispo}' — only INTERESTED or CALL BACK allowed"))
    return errors, warnings

def validate_dialer(rows):
    errors, warnings = [], []
    required = ["session_uuid", "init_time", "callmode", "talk_time", "user"]
    sheet = "Dialer Dump"
    if not rows:
        return [_e(sheet, "-", "Sheet is empty or not found")], []
    for col in required:
        if col not in rows[0]:
            errors.append(_e(sheet, "Header", f"Missing required column: '{col}'"))
    if errors:
        return errors, warnings
    for i, r in enumerate(rows, 2):
        uuid     = ss(r.get("session_uuid", ""))
        callmode = ss(r.get("callmode", "")).lower()
        talk     = sf(r.get("talk_time", 0))
        dispo    = ss(r.get("primary_dispo", ""))
        dialed   = ss(r.get("dialed_status", ""))
        date     = excel_to_date(r.get("init_time"))
        if not uuid:
            errors.append(_e(sheet, str(i), "session_uuid is missing"))
        if date is None:
            errors.append(_e(sheet, str(i), "Invalid or missing init_time"))
        if callmode and callmode not in VALID_CALLMODES:
            errors.append(_e(sheet, str(i), f"Unknown callmode '{callmode}'"))
        if talk < 0:
            errors.append(_e(sheet, str(i), f"talk_time is negative ({talk})"))
        if dialed.lower() == "connected" and not dispo:
            warnings.append(_w(sheet, str(i), "Connected call has no disposition"))
    return errors, warnings

def validate_pipeline(rows):
    errors, warnings = [], []
    required = ["Borrower Name", "Lender Name", "Status", "Date of Lead"]
    sheet = "Login Pipeline"
    if not rows:
        return [_e(sheet, "-", "Sheet is empty or not found")], []
    for col in required:
        if col not in rows[0]:
            errors.append(_e(sheet, "Header", f"Missing required column: '{col}'"))
    if errors:
        return errors, warnings
    shifted_col_hits = 0
    for i, r in enumerate(rows, 2):
        borrower = ss(r.get("Borrower Name", ""))
        lender   = ss(r.get("Lender Name", ""))
        amt      = r.get("Expected Sanction Amount(cr)")
        date     = excel_to_date(r.get("Date of Lead"))
        officer  = ss(r.get("Officer Name", ""))
        if not borrower:
            errors.append(_e(sheet, str(i), "Borrower Name is missing"))
        if not lender:
            errors.append(_e(sheet, str(i), "Lender Name is missing"))
        if date is None:
            errors.append(_e(sheet, str(i),
                f"Invalid Date of Lead for '{borrower or 'unknown'}'"))
        if amt is not None and sf(amt) < 0:
            errors.append(_e(sheet, str(i),
                f"Expected Sanction Amount is negative for '{borrower}'"))
        if not officer:
            warnings.append(_w(sheet, str(i),
                f"Officer Name missing for '{borrower}'"))

        # ── Status validation / column-shift guard ────────────────────────────
        # An extra or missing column shifts every field one position, so the
        # Status cell ends up holding the sanction amount or a date. Left
        # unchecked those values flow straight through to the dashboard as
        # phantom status cards, and the mis-read Lender Name silently
        # duplicates pipeline entries. Catch it at upload instead.
        raw = r.get("Status")
        if isinstance(raw, bool):
            errors.append(_e(sheet, str(i), f"Status is a boolean for '{borrower}'"))
        elif isinstance(raw, (int, float)):
            shifted_col_hits += 1
            errors.append(_e(sheet, str(i),
                f"Status contains a number ({raw}) for '{borrower}' — this is normally "
                f"the Expected Sanction Amount, which means a column is shifted"))
        elif isinstance(raw, (datetime.datetime, datetime.date)):
            shifted_col_hits += 1
            errors.append(_e(sheet, str(i),
                f"Status contains a date ({raw}) for '{borrower}' — a column is shifted"))
        else:
            s = ss(raw)
            if not s:
                warnings.append(_w(sheet, str(i),
                    f"Status is blank for '{borrower}' — will show as 'Unknown'"))
            else:
                _norm_s, _fuzzy = normalize_status(s)
                if _norm_s not in KNOWN_STATUSES:
                    errors.append(_e(sheet, str(i),
                        f"Unrecognised Status '{s}' for '{borrower}'. Allowed: "
                        + ", ".join(KNOWN_STATUSES)))

    if shifted_col_hits >= 3:
        errors.insert(0, _e(sheet, "Header",
            f"{shifted_col_hits} rows have a number or date in the Status column. "
            f"The sheet almost certainly has an extra or missing column (a City "
            f"column inserted after Borrower Name causes exactly this). Fix the "
            f"column alignment and re-upload."))
    return errors, warnings


# ── Aggregators ───────────────────────────────────────────────────────────────

def aggregate_ai(rows, processed_dates, existing_unique=None, existing_monthly_ai=None):
    by_day = {}
    added = skipped = 0
    new_unique = set(existing_unique) if existing_unique else set()
    new_monthly_ai = {}
    for ym, s in (existing_monthly_ai or {}).items():
        new_monthly_ai[ym] = set(s)

    for r in rows:
        date = excel_to_date(r.get("Call Date"))
        if not date:
            continue
        if date in processed_dates:
            skipped += 1
            continue
        dispo    = ss(r.get("L1 Disposition", "")).upper() or "BLANK"
        dur      = sf(r.get("Call Duration", 0))
        cust     = ss(r.get("Customer Number", ""))
        is_conn  = dispo not in NOT_CONNECTED_DISPOS
        if cust:
            new_unique.add(cust)
            ym = date[:7]
            if ym not in new_monthly_ai:
                new_monthly_ai[ym] = set()
            new_monthly_ai[ym].add(cust)

        if date not in by_day:
            by_day[date] = {"_leads": set(), "total_dials": 0, "connected": 0, "dispositions": {}}
        d = by_day[date]
        d["_leads"].add(cust)
        d["total_dials"] += 1
        if is_conn:
            d["connected"] += 1
        ds = d["dispositions"]
        if dispo not in ds:
            ds[dispo] = {"dials": 0, "connects": 0, "talk_time_sec": 0.0}
        ds[dispo]["dials"]         += 1
        ds[dispo]["connects"]      += (1 if is_conn else 0)
        ds[dispo]["talk_time_sec"] += dur
        added += 1

    result = {}
    for date, d in by_day.items():
        dispos = {}
        for k, v in d["dispositions"].items():
            dispos[k] = {**v,
                "aht_sec": v["talk_time_sec"] / v["connects"] if v["connects"] > 0 else 0.0}
        result[date] = {
            "unique_leads": len(d["_leads"]),
            "total_dials":  d["total_dials"],
            "connected":    d["connected"],
            "dispositions": dispos,
        }
    return result, set(result.keys()), added, skipped, new_unique, new_monthly_ai


def aggregate_handoff(rows, processed_dates, existing_handoff_mobiles=None):
    by_day = {}
    added = skipped = 0
    new_handoff_mobiles = set(existing_handoff_mobiles) if existing_handoff_mobiles else set()

    for r in rows:
        date  = excel_to_date(r.get("Call Date"))
        if not date:
            continue
        if date in processed_dates:
            skipped += 1
            continue
        dispo  = ss(r.get("L1 Disposition", "")).upper()
        batch  = ss(r.get("Batch", "Unknown")) or "Unknown"
        mobile = normalize_mobile(ss(r.get("Customer Number") or ""))
        if mobile:
            new_handoff_mobiles.add(mobile)

        if date not in by_day:
            by_day[date] = {"interested": 0, "callback": 0, "batches": {}}
        d = by_day[date]
        if dispo == "INTERESTED":
            d["interested"] += 1
        elif dispo == "CALL BACK":
            d["callback"] += 1
        if batch not in d["batches"]:
            d["batches"][batch] = {"interested": 0, "callback": 0}
        if dispo == "INTERESTED":
            d["batches"][batch]["interested"] += 1
        elif dispo == "CALL BACK":
            d["batches"][batch]["callback"]   += 1
        added += 1

    return by_day, set(by_day.keys()), added, skipped, new_handoff_mobiles


def aggregate_dialer(rows, processed_dates_by_source, handoff_set=None,
                     existing_blended_cids=None, existing_manual_cids=None,
                     existing_monthly_blen=None, existing_monthly_manual=None):
    """
    `processed_dates_by_source` is {"flexy": {dates...}, "jarvis": {dates...}}.

    Flexy and Jarvis are two independent feeds that arrive on different
    timelines, so each dedupes against its OWN processed-date set. Sharing one
    set let whichever feed landed first for a given date permanently block the
    other feed's data for that same date.
    """
    blended_by_day = {}
    manual_by_day  = {}
    added = skipped = discarded = 0
    dates_by_source   = {"flexy": set(), "jarvis": set()}
    skipped_by_source = {"flexy": 0,     "jarvis": 0}
    agent_sources     = {"user": 0, "supervisor_name": 0}
    new_blended_cids = set(existing_blended_cids) if existing_blended_cids else set()
    new_manual_cids  = set(existing_manual_cids)  if existing_manual_cids  else set()
    new_monthly_blen = {}
    for ym, s in (existing_monthly_blen or {}).items():
        new_monthly_blen[ym] = set(s)
    new_monthly_manual = {}
    for ym, s in (existing_monthly_manual or {}).items():
        new_monthly_manual[ym] = set(s)
    h_set = handoff_set or set()

    for r in rows:
        source    = r.get("_source", "flexy")
        pd_source = processed_dates_by_source.get(source, set())
        date = excel_to_date(r.get("init_time"))
        if not date:
            continue
        if date in pd_source:
            skipped += 1
            skipped_by_source[source] = skipped_by_source.get(source, 0) + 1
            continue

        cid      = normalize_mobile(ss(r.get("customer_cid") or ""))
        # Agent identity: Flexy exports leave `user` blank and carry the real
        # agent in `supervisor_name`. Rows with neither are the predictive
        # dialer's raw attempt legs — their successful outcomes are logged
        # separately as agent rows, so discarding them here is correct and
        # avoids double-counting dials.
        agent    = ss(r.get("user", ""))
        if agent:
            agent_sources["user"] += 1
        else:
            agent = ss(r.get("supervisor_name", ""))
            if agent:
                agent_sources["supervisor_name"] += 1
        if not agent:
            discarded += 1
            continue
        talk_sec = sf(r.get("talk_time", 0)) * 86400.0
        dispo    = ss(r.get("primary_dispo", "")) or "No Feedback"
        is_conn  = ss(r.get("dialed_status", "")).lower() == "connected"

        if cid and cid in h_set:
            target = blended_by_day
            new_blended_cids.add(cid)
            ym = date[:7]
            if ym not in new_monthly_blen:
                new_monthly_blen[ym] = set()
            new_monthly_blen[ym].add(cid)
        else:
            target = manual_by_day
            if cid:
                new_manual_cids.add(cid)
                ym = date[:7]
                if ym not in new_monthly_manual:
                    new_monthly_manual[ym] = set()
                new_monthly_manual[ym].add(cid)

        if date not in target:
            target[date] = {"_unique": set(), "total_dials": 0, "connected": 0,
                            "talk_time_sec": 0.0, "dispositions": {}, "agents": {}}
        d = target[date]
        if cid: d["_unique"].add(cid)
        d["total_dials"]   += 1
        d["connected"]     += (1 if is_conn else 0)
        d["talk_time_sec"] += talk_sec
        d["dispositions"][dispo] = d["dispositions"].get(dispo, 0) + 1

        if agent not in d["agents"]:
            d["agents"][agent] = {"_unique": set(), "dials": 0, "connected": 0,
                                   "talk_time_sec": 0.0, "dispositions": {}}
        a = d["agents"][agent]
        if cid: a["_unique"].add(cid)
        a["dials"]        += 1
        a["connected"]    += (1 if is_conn else 0)
        a["talk_time_sec"]+= talk_sec
        a["dispositions"][dispo] = a["dispositions"].get(dispo, 0) + 1
        added += 1
        dates_by_source[source].add(date)

    def finalize(by_day_dict):
        result = {}
        for date, d in by_day_dict.items():
            agents_out = {}
            for ag, av in d["agents"].items():
                agents_out[ag] = {
                    "unique_leads":  len(av["_unique"]),
                    "dials":         av["dials"],
                    "connected":     av["connected"],
                    "talk_time_sec": av["talk_time_sec"],
                    "dispositions":  av["dispositions"],
                }
            result[date] = {
                "unique_leads":  len(d["_unique"]),
                "total_dials":   d["total_dials"],
                "connected":     d["connected"],
                "talk_time_sec": d["talk_time_sec"],
                "dispositions":  d["dispositions"],
                "agents":        agents_out,
            }
        return result

    stats = {
        "added": added, "skipped": skipped, "discarded": discarded,
        "skipped_by_source": skipped_by_source, "agent_sources": agent_sources,
    }
    return (finalize(blended_by_day), finalize(manual_by_day),
            dates_by_source, stats, new_blended_cids, new_manual_cids,
            new_monthly_blen, new_monthly_manual)


def merge_dialer_by_day(stored_by_day, new_by_day):
    """
    Additively merge new dialer day-stats into stored ones.

    dict.update() would REPLACE a date's entry. That was harmless while Flexy
    and Jarvis shared one processed-date set (only one feed could ever own a
    given date), but once each feed dedupes independently both can contribute
    to the same calendar day across different uploads — and a plain update()
    silently destroys whichever feed landed first.
    """
    for date, new_day in new_by_day.items():
        cur = stored_by_day.get(date)
        if not cur:
            stored_by_day[date] = new_day
            continue
        cur["unique_leads"]  = cur.get("unique_leads", 0)  + new_day.get("unique_leads", 0)
        cur["total_dials"]   = cur.get("total_dials", 0)   + new_day.get("total_dials", 0)
        cur["connected"]     = cur.get("connected", 0)     + new_day.get("connected", 0)
        cur["talk_time_sec"] = cur.get("talk_time_sec", 0.0) + new_day.get("talk_time_sec", 0.0)

        cur_d = cur.setdefault("dispositions", {})
        for dispo, cnt in new_day.get("dispositions", {}).items():
            cur_d[dispo] = cur_d.get(dispo, 0) + cnt

        cur_a = cur.setdefault("agents", {})
        for agent, s in new_day.get("agents", {}).items():
            a = cur_a.get(agent)
            if not a:
                cur_a[agent] = s
                continue
            a["unique_leads"]  = a.get("unique_leads", 0)  + s.get("unique_leads", 0)
            a["dials"]         = a.get("dials", 0)         + s.get("dials", 0)
            a["connected"]     = a.get("connected", 0)     + s.get("connected", 0)
            a["talk_time_sec"] = a.get("talk_time_sec", 0.0) + s.get("talk_time_sec", 0.0)
            ad = a.setdefault("dispositions", {})
            for dispo, cnt in s.get("dispositions", {}).items():
                ad[dispo] = ad.get(dispo, 0) + cnt
    return stored_by_day


def process_pipeline(rows, existing_rows):
    """
    The uploaded Login Pipeline sheet is the SINGLE SOURCE OF TRUTH.

    Each upload carries the complete current pipeline, so the stored pipeline is
    rebuilt from the sheet rather than merged into. Merging meant every
    borrower/lender pair ever seen stayed forever: pairs that were renamed,
    corrected or removed in later sheets lingered and inflated the entry count
    (56 real entries were showing as 73 before this changed).
    """
    existing = {}
    for row in existing_rows:
        existing[f"{row.get('borrower','')}|{row.get('lender','')}"] = row

    pipeline = {}
    added = updated = 0
    fuzzy_warns = []

    for r in rows:
        borrower = ss(r.get("Borrower Name", ""))
        lender   = ss(r.get("Lender Name", ""))
        if not borrower or not lender:
            continue

        date_lead  = excel_to_date(r.get("Date of Lead"))
        label_lead = date_label(date_lead) if date_lead else ""
        key        = f"{borrower}|{lender}"

        status_raw  = ss(r.get("Status", ""))
        status_norm, was_fuzzy = normalize_status(status_raw)
        if was_fuzzy:
            fuzzy_warns.append(
                f"Status '{status_raw}' matched to '{status_norm}' for {borrower} / {lender}")

        row_data = {
            "date_of_lead":          label_lead,
            "officer":               ss(r.get("Officer Name", "")),
            "borrower":              borrower,
            "mobile":                ss(r.get("Mobile No.", "")),
            "dsa_code":              ss(r.get("DSA Code", "")),
            "lender":                lender,
            "lead_source":           ss(r.get("Lead Source", "")),
            "lead_type":             ss(r.get("Lead Type", "")),
            "product_type":          ss(r.get("Product Type", "")),
            "login_date":            date_label(excel_to_date(r.get("Login Date"))),
            "target_sanction_date":  date_label(excel_to_date(r.get("Target Sanction Date"))),
            "target_disbursal_date": date_label(excel_to_date(r.get("Target Disbursement Date"))),
            "expected_amount_cr":    sf(r.get("Expected Sanction Amount(cr)", 0)),
            "status":                status_norm,
            "status_raw":            status_raw,
            "remarks":               ss(r.get("Remarks", "")),
        }

        if key in existing:
            updated += 1
        else:
            added += 1
        pipeline[key] = row_data

    dropped = [k for k in existing if k not in pipeline]
    return list(pipeline.values()), added, updated, dropped, fuzzy_warns


# ── Response builder ──────────────────────────────────────────────────────────

def _sum_ai(by_day, dates):
    r = {"unique_leads": 0, "total_dials": 0, "connected": 0, "dispositions": {}}
    for d in dates:
        day = by_day.get(d, {})
        r["unique_leads"] += day.get("unique_leads", 0)
        r["total_dials"]  += day.get("total_dials", 0)
        r["connected"]    += day.get("connected", 0)
        for dispo, s in day.get("dispositions", {}).items():
            if dispo not in r["dispositions"]:
                r["dispositions"][dispo] = {"dials": 0, "connects": 0, "talk_time_sec": 0.0}
            r["dispositions"][dispo]["dials"]         += s.get("dials", 0)
            r["dispositions"][dispo]["connects"]      += s.get("connects", 0)
            r["dispositions"][dispo]["talk_time_sec"] += s.get("talk_time_sec", 0.0)
    for dispo in r["dispositions"]:
        s = r["dispositions"][dispo]
        s["aht_sec"] = s["talk_time_sec"] / s["connects"] if s["connects"] > 0 else 0.0
    return r

def _sum_handoff(by_day, dates):
    r = {"interested": 0, "callback": 0, "batches": {}}
    for d in dates:
        day = by_day.get(d, {})
        r["interested"] += day.get("interested", 0)
        r["callback"]   += day.get("callback", 0)
        for b, s in day.get("batches", {}).items():
            if b not in r["batches"]:
                r["batches"][b] = {"interested": 0, "callback": 0}
            r["batches"][b]["interested"] += s.get("interested", 0)
            r["batches"][b]["callback"]   += s.get("callback", 0)
    return r

def _sum_dialer(by_day, dates):
    r = {"unique_leads": 0, "total_dials": 0, "connected": 0, "talk_time_sec": 0.0,
         "dispositions": {}, "agents": {}}
    for d in dates:
        day = by_day.get(d, {})
        r["unique_leads"]  += day.get("unique_leads", 0)
        r["total_dials"]   += day.get("total_dials", 0)
        r["connected"]     += day.get("connected", 0)
        r["talk_time_sec"] += day.get("talk_time_sec", 0.0)
        for dispo, cnt in day.get("dispositions", {}).items():
            r["dispositions"][dispo] = r["dispositions"].get(dispo, 0) + cnt
        for agent, s in day.get("agents", {}).items():
            if agent not in r["agents"]:
                r["agents"][agent] = {
                    "unique_leads": 0, "dials": 0, "connected": 0,
                    "talk_time_sec": 0.0, "dispositions": {}
                }
            a = r["agents"][agent]
            a["unique_leads"] += s.get("unique_leads", 0)
            a["dials"]        += s.get("dials", 0)
            a["connected"]    += s.get("connected", 0)
            a["talk_time_sec"]+= s.get("talk_time_sec", 0.0)
            for dispo, cnt in s.get("dispositions", {}).items():
                a["dispositions"][dispo] = a["dispositions"].get(dispo, 0) + cnt
    return r

def build_response(stored):
    ai_dates  = sorted(stored.get("ai_by_day", {}).keys())
    h_dates   = sorted(stored.get("handoff_by_day", {}).keys())
    b_dates   = sorted(stored.get("dialer_blended_by_day", {}).keys())
    dm_dates  = sorted(stored.get("dialer_manual_by_day", {}).keys())

    def package(sum_fn, by_day, dates):
        mtd = sum_fn(by_day, dates)
        mtd["label"] = "MTD"
        last = None
        if dates:
            last = sum_fn(by_day, [dates[-1]])
            last["label"] = date_label(dates[-1])
        all_days = []
        for d in dates:
            day = sum_fn(by_day, [d])
            day["label"] = date_label(d)
            all_days.append(day)
        return {"mtd": mtd, "lastDay": last, "allDays": all_days}

    handoff_set  = load_set("handoff_mobile_set")
    blended_set  = load_set("blended_cid_set")
    not_dialled  = len(handoff_set - blended_set)

    return {
        "generated":            stored.get("generated", ""),
        "uploadedBy":           stored.get("uploadedBy", ""),
        "totalUniqueLeads":     stored.get("total_unique_leads", 0),
        "totalBlenUnique":      stored.get("total_blended_unique", 0),
        "totalManualUnique":    stored.get("total_manual_unique", 0),
        "handoffNotDialled":    not_dialled,
        "totalHandoffUnique":   len(handoff_set),
        "monthlyUniqueAI":      db_get("monthly_unique_counts", {}).get("ai",     {}),
        "monthlyUniqueBlen":    db_get("monthly_unique_counts", {}).get("blen",   {}),
        "monthlyUniqueManual":  db_get("monthly_unique_counts", {}).get("manual", {}),
        "ai":           package(_sum_ai,      stored.get("ai_by_day", {}),              ai_dates),
        "handoff":      package(_sum_handoff, stored.get("handoff_by_day", {}),         h_dates),
        "dialerBlended":package(_sum_dialer,  stored.get("dialer_blended_by_day", {}),  b_dates),
        "dialerManual": package(_sum_dialer,  stored.get("dialer_manual_by_day", {}),   dm_dates),
        "pipeline":     {"rows": stored.get("pipeline_rows", [])},
    }


# ── Audit ─────────────────────────────────────────────────────────────────────

def append_audit(entry):
    log = db_get("audit_log") or []
    log.insert(0, entry)
    db_set("audit_log", log[:100])


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def home():
    return "CarmaOne MIS Backend is running"


@app.route("/upload", methods=["POST"])
def upload_file():
    pin = request.form.get("pin", "").strip()
    if not pin:
        return jsonify({"error": "PIN is required"}), 403
    uploader = PINS.get(pin)
    if not uploader:
        return jsonify({"error": "Invalid PIN. Please check and try again."}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files allowed"}), 400

    tmp = "upload_tmp.xlsx"
    file.save(tmp)

    try:
        wb = load_workbook(tmp, read_only=True, data_only=True)

        # Locate sheets — supports Flexy_Dump, Dialer_Dump, Jarvis_Dump
        sn_ai     = find_sheet_name(wb, ["AI_Dump", "AI Dump", "AIDump"])
        sn_hand   = find_sheet_name(wb, ["Handoff_Leads", "Handoff Leads", "Handoff"])
        sn_dial   = find_sheet_name(wb, ["Flexy_Dump", "Flexy Dump", "Dialer_Dump", "Dialer Dump", "DialerDump", "Flexy"])
        sn_jarvis = find_sheet_name(wb, ["Jarvis_Dump", "Jarvis Dump", "Jarvis"])
        sn_pipe   = find_sheet_name(wb, ["Login_Pipeline", "Login Pipeline", "Pipeline"])

        missing_hard = [n for n, s in [
            ("AI_Dump", sn_ai), ("Handoff_Leads", sn_hand)
        ] if not s]
        if not sn_dial and not sn_jarvis:
            missing_hard.append("Dialer/Flexy/Jarvis dump")

        if missing_hard:
            wb.close(); os.remove(tmp)
            return jsonify({"error": f"Missing required sheets: {', '.join(missing_hard)}. "
                            f"Found: {', '.join(wb.sheetnames)}"}), 422

        pipeline_missing = not sn_pipe

        # Load smaller sheets normally
        h_rows      = read_sheet(wb, sn_hand)
        d_rows      = read_sheet(wb, sn_dial)      if sn_dial   else []
        jarvis_rows = read_sheet(wb, sn_jarvis)    if sn_jarvis else []
        p_rows      = read_sheet(wb, sn_pipe)      if sn_pipe   else []

        # Validate non-AI sheets before streaming AI
        all_errors, all_warnings = [], []
        for errs, warns in [
            validate_handoff(h_rows),
            validate_dialer(d_rows) if d_rows else ([], []),
        ]:
            all_errors.extend(errs)
            all_warnings.extend(warns)

        if p_rows:
            errs, warns = validate_pipeline(p_rows)
            all_errors.extend(errs)
            all_warnings.extend(warns)
        elif pipeline_missing:
            all_warnings.append(_w("Login Pipeline", "-",
                "Sheet not found in this upload — pipeline data unchanged"))

        if all_errors:
            wb.close(); os.remove(tmp)
            append_audit({
                "ts": now_ist(), "by": uploader, "status": "REJECTED",
                "errors":   [f"[{e['sheet']}] Row {e['row']}: {e['msg']}" for e in all_errors],
                "warnings": [], "changes": {}
            })
            return jsonify({
                "success": False,
                "validationErrors": all_errors,
                "message": "Upload rejected due to data errors. See details below."
            }), 422

        # Stream AI_Dump — never loads all rows into memory
        ai_stream = stream_sheet(wb, sn_ai,
                                 cols=["Customer Number","Call Date","L1 Disposition","Call Duration","Batch"]) \
                    if sn_ai else iter([])

        # Load existing stored data
        stored = db_get("mis_data") or {
            "processed_dates":         {"ai_dump": [], "handoff": [],
                                        "dialer_flexy": [], "dialer_jarvis": []},
            "pipeline_rows":           [],
            "ai_by_day":               {},
            "handoff_by_day":          {},
            "dialer_blended_by_day":   {},
            "dialer_manual_by_day":    {},
        }

        if "dialer_manual" in stored.get("processed_dates", {}):
            stored["processed_dates"]["dialer"] = stored["processed_dates"].pop("dialer_manual", [])

        # Migration: force full reprocess when Jarvis data first introduced
        if not stored.get("jarvis_integrated"):
            stored["processed_dates"]["dialer"]  = []
            stored["processed_dates"]["handoff"] = []
            stored["dialer_manual_by_day"]   = {}
            stored["dialer_blended_by_day"]  = {}
            stored["handoff_by_day"]         = {}
            stored["jarvis_integrated"]      = True

        # Split the legacy shared "dialer" processed-date bucket into per-source
        # buckets. The shared bucket let Jarvis dates block Flexy data (and vice
        # versa) for the same calendar day, so any state built under it is
        # untrustworthy and gets rebuilt from source.
        _pd = stored["processed_dates"]
        if "dialer" in _pd:
            _pd.pop("dialer", None)
            _pd["dialer_flexy"]  = []
            _pd["dialer_jarvis"] = []
            stored["dialer_blended_by_day"] = {}
            stored["dialer_manual_by_day"]  = {}

        pd_ai       = set(_pd.get("ai_dump", []))
        pd_h        = set(_pd.get("handoff", []))
        pd_d_flexy  = set(_pd.get("dialer_flexy", []))
        pd_d_jarvis = set(_pd.get("dialer_jarvis", []))

        # Aggregate AI (streaming)
        existing_unique     = load_unique_set()
        existing_monthly_ai = db_get_monthly_sets("monthly_ai")
        ai_new, ai_dates, ai_add, ai_skip, new_unique, new_monthly_ai = aggregate_ai(
            ai_stream, pd_ai, existing_unique, existing_monthly_ai)
        wb.close(); os.remove(tmp); gc.collect()

        # Aggregate Handoff
        existing_handoff_mobiles = load_set("handoff_mobile_set")
        h_new, h_dates, h_add, h_skip, new_handoff_mobiles = aggregate_handoff(
            h_rows, pd_h, existing_handoff_mobiles)

        # Merge Jarvis into Flexy, normalize agent names, tag each row's source
        for r in d_rows:
            if r.get("user"):
                r["user"] = normalize_agent(ss(r["user"]))
            r["_source"] = "flexy"
        jarvis_normalized = normalize_jarvis_rows(jarvis_rows)  # tagged "jarvis"
        combined_d_rows   = d_rows + jarvis_normalized
        del jarvis_rows; gc.collect()

        # Aggregate Dialer
        handoff_set            = new_handoff_mobiles
        existing_blended_cids  = load_set("blended_cid_set")
        existing_manual_cids   = load_set("manual_cid_set")
        existing_monthly_blen  = db_get_monthly_sets("monthly_blen")
        existing_monthly_manual= db_get_monthly_sets("monthly_manual")
        b_new, dm_new, d_dates_by_source, d_stats, new_blended_cids, new_manual_cids, \
            new_monthly_blen, new_monthly_manual = aggregate_dialer(
            combined_d_rows, {"flexy": pd_d_flexy, "jarvis": pd_d_jarvis}, handoff_set,
            existing_blended_cids, existing_manual_cids,
            existing_monthly_blen, existing_monthly_manual)
        del combined_d_rows, d_rows; gc.collect()

        # Pipeline
        p_list, p_add, p_upd, p_dropped, p_warns = process_pipeline(
            p_rows, stored.get("pipeline_rows", []))

        # Merge & save
        stored["ai_by_day"].update(ai_new)
        save_unique_set(new_unique)
        stored["total_unique_leads"] = len(new_unique)

        stored["handoff_by_day"].update(h_new)
        save_set("handoff_mobile_set", new_handoff_mobiles)

        merge_dialer_by_day(stored["dialer_blended_by_day"], b_new)
        merge_dialer_by_day(stored["dialer_manual_by_day"],  dm_new)
        save_set("blended_cid_set", new_blended_cids)
        save_set("manual_cid_set",  new_manual_cids)
        stored["total_blended_unique"] = len(new_blended_cids)
        stored["total_manual_unique"]  = len(new_manual_cids)

        db_set_monthly_sets("monthly_ai",     new_monthly_ai)
        db_set_monthly_sets("monthly_blen",   new_monthly_blen)
        db_set_monthly_sets("monthly_manual", new_monthly_manual)
        db_set("monthly_unique_counts", {
            "ai":     {ym: len(s) for ym, s in new_monthly_ai.items()},
            "blen":   {ym: len(s) for ym, s in new_monthly_blen.items()},
            "manual": {ym: len(s) for ym, s in new_monthly_manual.items()},
        })

        if p_rows:
            stored["pipeline_rows"] = p_list
        stored["processed_dates"]["ai_dump"]       = sorted(pd_ai | ai_dates)
        stored["processed_dates"]["handoff"]       = sorted(pd_h  | h_dates)
        stored["processed_dates"]["dialer_flexy"]  = sorted(pd_d_flexy  | d_dates_by_source["flexy"])
        stored["processed_dates"]["dialer_jarvis"] = sorted(pd_d_jarvis | d_dates_by_source["jarvis"])
        stored["generated"]        = now_ist()
        stored["uploadedBy"]       = uploader
        stored["last_dialer_stats"] = d_stats
        db_set("mis_data", stored)

        changes = {
            "AI Dump":        f"{ai_add} rows added, {ai_skip} skipped (already processed)",
            "Handoff Leads":  f"{h_add} rows added, {h_skip} skipped",
            "Dialer Dump":    (f"{d_stats['added']} calls counted "
                               f"({len(b_new)} blended days, {len(dm_new)} direct days); "
                               f"{d_stats['skipped']} skipped as already-processed "
                               f"(flexy {d_stats['skipped_by_source'].get('flexy',0)}, "
                               f"jarvis {d_stats['skipped_by_source'].get('jarvis',0)}); "
                               f"{d_stats['discarded']} discarded — no agent identity"),
            "Login Pipeline": (f"{len(p_list)} entries now live "
                               f"({p_add} new, {p_upd} updated, "
                               f"{len(p_dropped)} no longer in sheet — removed)"
                               if p_rows else "Not in this upload — unchanged"),
        }

        warn_strs = ([f"[{w['sheet']}] Row {w['row']}: {w['msg']}" for w in all_warnings]
                     + [f"[Login Pipeline] {w}" for w in p_warns]
                     + ([f"[Login Pipeline] {len(p_dropped)} entr"
                         f"{'y' if len(p_dropped)==1 else 'ies'} removed — no longer "
                         f"present in the uploaded sheet: "
                         + ", ".join(k.replace('|', ' / ') for k in p_dropped[:5])
                         + (" …" if len(p_dropped) > 5 else "")]
                        if p_dropped else []))

        append_audit({
            "ts": now_ist(), "by": uploader, "status": "SUCCESS",
            "errors": [], "warnings": warn_strs, "changes": changes
        })

        return jsonify({
            "success":  True,
            "message":  f"File uploaded successfully by {uploader}",
            "changes":  changes,
            "warnings": warn_strs,
        })

    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return jsonify({"error": str(e)}), 500


@app.route("/data", methods=["GET"])
def get_data():
    stored = db_get("mis_data")
    if not stored:
        return jsonify({"error": "No data available"}), 404
    resp = make_response(jsonify(build_response(stored)))
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.route("/audit", methods=["GET"])
def get_audit():
    log = db_get("audit_log") or []
    resp = make_response(jsonify(log))
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.route("/status", methods=["GET"])
def get_status():
    stored = db_get("mis_data")
    if not stored:
        return jsonify({"uploaded": False, "message": "No data uploaded yet"})
    return jsonify({
        "uploaded":   True,
        "generated":  stored.get("generated", ""),
        "uploadedBy": stored.get("uploadedBy", ""),
    })


@app.route("/dbcheck")
def dbcheck():
    conn = get_db()
    if conn is None:
        return jsonify({
            "connected": False,
            "error": "Cannot connect to database. Check DB_HOST, DB_USER, DB_PORT env vars.",
            "db_host": os.environ.get("DB_HOST", "NOT SET"),
            "db_user": os.environ.get("DB_USER", "NOT SET"),
            "db_port": os.environ.get("DB_PORT", "NOT SET"),
        })
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM carmaone_kv")
            row_count = cur.fetchone()[0]
        return jsonify({
            "connected": True,
            "table_exists": True,
            "row_count": row_count,
            "db_host": os.environ.get("DB_HOST", "NOT SET"),
            "db_user": os.environ.get("DB_USER", "NOT SET"),
            "db_port": os.environ.get("DB_PORT", "NOT SET"),
        })
    except Exception as e:
        return jsonify({"connected": True, "table_exists": False, "error": str(e)})


@app.route("/debug")
def debug():
    handoff_set = load_set("handoff_mobile_set")
    blended_set = load_set("blended_cid_set")
    manual_set  = load_set("manual_cid_set")
    stored      = db_get("mis_data") or {}
    return jsonify({
        "handoff_mobile_count":    len(handoff_set),
        "blended_cid_count":       len(blended_set),
        "manual_cid_count":        len(manual_set),
        "handoff_sample":          list(handoff_set)[:10],
        "blended_sample":          list(blended_set)[:10],
        "dialer_blended_days":     list(stored.get("dialer_blended_by_day", {}).keys()),
        "dialer_manual_day_count": len(stored.get("dialer_manual_by_day", {})),
        "processed_dates":         stored.get("processed_dates", {}),
        "jarvis_integrated":       stored.get("jarvis_integrated", False),
        "last_dialer_stats":       stored.get("last_dialer_stats", {}),
    })

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
